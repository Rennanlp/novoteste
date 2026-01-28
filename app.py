# app.py
from flask import Flask, render_template, request, send_file, redirect, session, url_for, jsonify, g, render_template_string, json, make_response, flash
from unidecode import unidecode
import os
import csv
from datetime import datetime, timedelta
from functools import wraps
import xlsxwriter
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_migrate import Migrate
from collections import defaultdict
from jinja2 import Environment
import requests
import sqlite3
from werkzeug.utils import secure_filename
import io
import pandas as pd
from io import BytesIO
import asyncio
import aiohttp
from clientes import criar_banco_dados, inserir_dados_da_planilha, obter_responsaveis_e_empresas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Frame, PageTemplate
from reportlab.pdfgen import canvas
import io
import textwrap
from flask_mail import Mail, Message
from openpyxl import Workbook
from sqlalchemy import or_
import pytz
import pymysql
import boto3
from botocore.exceptions import NoCredentialsError, ClientError
import re
from dotenv import load_dotenv
import logging
from api_client import APIRelatorioReversoClient
from difflib import SequenceMatcher
from learning import LearningSystem

# CONFUGURAÇÕES FLASK #

app = Flask(__name__)
load_dotenv()
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or b'_5#y2L"F4Q8z\n\xec]/'
app.config['UPLOAD_FOLDER'] = 'static/images'
app.config['UPLOAD_FOLDER1'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'csv', }
app.config['STATIC_FOLDER'] = 'static'
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
app.config['AWS_SECRET_ACCESS_KEY'] = os.getenv('AWS_SECRET_ACCESS_KEY')
app.config['AWS_ACCESS_KEY_ID'] = os.getenv('AWS_ACCESS_KEY_ID')
app.config['AWS_S3_BUCKET_NAME'] = os.getenv('AWS_S3_BUCKET_NAME')
app.config['AWS_S3_REGION_NAME'] = os.getenv('AWS_S3_REGION_NAME', 'us-west-2')
app.config['MAX_CONTENT_LENGTH'] = 48 * 1024 * 1024
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_BINDS'] = {
    'database1': 'sqlite:///database1.db'
}
db = SQLAlchemy(app)
CORS(app)
migrate = Migrate(app, db)

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Instância do cliente da API
api_client = APIRelatorioReversoClient()


# CONFIGURAÇÕES MAIL #
app.config['MAIL_SERVER'] = 'smtp.sendgrid.net'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'apikey' 
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = 'conexaopremium26@gmail.com'


mail = Mail(app)


def format_date(value, format='%d/%m/%Y'):

    if isinstance(value, str):
        return value

    return value.strftime(format)

app.jinja_env.filters['date'] = format_date

user_tasks = {}
user_quantities = {}
saved_data = {}
dashboard_data = {}

# banco de usuários
user_database = {
    'conexao.premium': {'password': 'senha123', 'name': 'Conexão Premium', 'email': 'conexao.premium@example.com'},
    'Renan10': {'password': '162593', 'name': 'Renan', 'email': 'rennanlpbass@gmail.com'},
    'Thais10': {'password': 'arthur08', 'name': 'Thais', 'email': 'thais10@example.com'},
    'Diogo10': {'password': '15222431', 'name': 'Diogo', 'email': 'diogo10@example.com'},
    'Keyse10': {'password': 'keyse321', 'name': 'Keyse', 'email': 'keyse10@example.com'},
    'Nicoli10': {'password': 'ririserelepe', 'name': 'Nic', 'email': 'nicoli10@example.com'},
    'Ingrid10': {'password': '135781', 'name': 'Ingrid', 'email': 'ingrid10@example.com'},
    'Eduarda': {'password': 'eduarda10', 'name': 'Duda', 'email': 'eduarda@example.com'},
    'Felipe': {'password': 'felipe@10', 'name': 'Felipe', 'email': 'felipe@example.com'},
    'Gabriela': {'password': 'gabriela@10', 'name': 'Gabriela', 'email': 'gabriela@example.com'},
    'Kymberli': {'password': 'kym@10', 'name': 'Kym', 'email': 'kymberli@example.com'},
    'Ana Paula': {'password': 'ana@10', 'name': 'Ana', 'email': 'ana.paula@example.com'},
    'Juliana': {'password': 'juliana@10', 'name': 'Juliana', 'email': 'juliana@example.com'},
    'ConexaoCD': {'password': 'conexaocd@10', 'name': 'Conexão CD', 'email': 'conexaocd@example.com'}
}

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), nullable=False)
    date = db.Column(db.String(10), nullable=False)
    content = db.Column(db.Text, nullable=False)

def login_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return view(*args, **kwargs)
    return wrapped_view

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def generate_excel(username, tasks):
    print("User tasks:", user_tasks)
    
    # Obtem os dados do formulário
    data = request.form.get('data')
    observations = request.form.getlist('observations[]')
    qtd = request.form.getlist('number1[]')

    print("Tasks:", tasks)
    print("Data:", data)
    print("Observations:", observations)
    print("Quantidades:", qtd)

    output_filepath = os.path.join(app.config['UPLOAD_FOLDER1'], 'lista_de_tarefas.xlsx')

    # Gera arquivo Excel
    workbook = xlsxwriter.Workbook(output_filepath)
    worksheet = workbook.add_worksheet()

    # cabeçalhos
    worksheet.write(0, 0, 'Tarefa')
    worksheet.write(0, 1, 'Data')
    worksheet.write(0, 2, 'Pedidos no Dia' )
    worksheet.write(0, 3, 'Observação')

    # Escreve as tarefas
    for i, task in enumerate(tasks, start=1):
        worksheet.write(i, 0, task)
        worksheet.write(i, 1, data)
        worksheet.write(i, 2, qtd[i - 1] if i <= len(qtd) else '')
        if i <= len(observations):
            worksheet.write(i, 3, observations[i - 1])


    workbook.close()

    return output_filepath

def process_csv(file):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    
    try:
        csv_reader = csv.reader(io.TextIOWrapper(file, encoding='latin-1'), delimiter=';')
    except Exception as e:
        print("Erro ao tentar abrir o arquivo CSV:", e)
        return

    next(csv_reader, None)

    for row in csv_reader:
        if len(row) >= 2 and all(row[:2]):
            id, cliente, token = row[0], row[1], row[2]
            c.execute("INSERT INTO clientes (id, cliente, token) VALUES (?, ?, ?)", (id, cliente, token))
        else:
            print("Ignorando linha do arquivo CSV com formato inválido:", row)

    conn.commit()
    conn.close()

def criar_tabela():
    conn = sqlite3.connect('clientes.db')
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS clientes (
                        id INTEGER PRIMARY KEY,
                        cliente TEXT,
                        token TEXT
                    )''')
    conn.commit()
    conn.close()

criar_tabela()

def obter_token_por_cliente(nome_cliente):
    conn = sqlite3.connect('clientes.db')
    cursor = conn.cursor()

    cursor.execute("SELECT token FROM clientes WHERE LOWER(cliente) = LOWER(?)", (nome_cliente,))
    resultado = cursor.fetchone()

    conn.close()

    if resultado:
        return resultado[0]
    else:
        return None

@app.route('/removedor')
@login_required
def removedor():
    username = user_database.get(session.get('username', ''), {}).get('name', 'Convidado')
    print("Username in session:", session.get('username'))
    return render_template('index.html', username=username)

@app.route('/login')
def login1():
    if 'username' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

# autenticar o login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['senha']

        if username in user_database and user_database[username]['password'] == password:
            session['username'] = username
            session['name'] = user_database[username]['name']
            return redirect(url_for('dashboard'))
        else:
            return "Credenciais inválidas", 401
    return render_template('login.html')

@app.route('/remove_accent', methods=['POST'])
@login_required
def remove_accent():
    if 'file' not in request.files:
        return "Nenhum arquivo enviado."

    file = request.files['file']

    if file.filename == '':
        return "Nenhum arquivo selecionado."

    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)

            encoding = request.form.get('encoding', 'utf-8')

            # Tenta abrir o arquivo CSV com o encoding fornecido
            try:
                with open(filepath, 'r', encoding=encoding) as input_file:
                    delimiter = ';'
                    reader = csv.reader(input_file, delimiter=delimiter)
                    rows = [list(map(lambda x: unidecode(x) if x else x, row)) for row in reader]
            except UnicodeDecodeError:
                # Se ocorrer um erro de decodificação, tentar abrir com 'latin-1'
                with open(filepath, 'r', encoding='latin-1') as input_file:
                    delimiter = ';'
                    reader = csv.reader(input_file, delimiter=delimiter)
                    rows = [list(map(lambda x: unidecode(x) if x else x, row)) for row in reader]

            # Criar um arquivo de saída para o novo CSV
            output_filepath = os.path.join(app.config['UPLOAD_FOLDER1'], f'Arquivo_Ajustado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
            with open(output_filepath, 'w', encoding='utf-8', newline='') as output_file:
                writer = csv.writer(output_file, delimiter=delimiter)
                writer.writerows(rows)

            return send_file(output_filepath, download_name=filename, as_attachment=True)

        except Exception as e:
            return "Erro durante o processamento do arquivo: {}".format(str(e))

    return "Tipo de arquivo não permitido."


@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/task')
@login_required
def task_f():
    username = session['username']
    user_task_list = user_tasks.get(username, [])
    return render_template('form.html', tasks=user_task_list)

@app.route('/add_task', methods=['POST'])
@login_required
def add_task():
    new_task = request.form.get('task')
    username = session['username']

    if username not in user_tasks:
        user_tasks[username] = []

    user_tasks[username].append(new_task)
    return redirect(url_for('task_f'))

@app.route('/download_excel', methods=['GET','POST'])
@login_required
def download_excel():
    username = session['username']
    tasks = user_tasks.get(username, [])
    
    # Gera o arquivo Excel usando o nome de usuário da sessão e as tarefas
    excel_filepath = generate_excel(username, tasks)

    # Envia o arquivo para download
    return send_file(excel_filepath, as_attachment=True)

@app.route('/remove_task', methods=['POST'])
@login_required
def remove_task():
    username = session['username']

    try:
        task_index = int(request.form.get('task_index'))

        if 0 <= task_index < len(user_tasks.get(username, [])):
            removed_task = user_tasks[username].pop(task_index)
            return jsonify({'status': 'success', 'removed_task': removed_task})
        else:
            return jsonify({'status': 'error', 'message': 'Índice de tarefa inválido'})
    except ValueError:
        return jsonify({'status': 'error', 'message': 'Índice de tarefa inválido'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/clear_tasks', methods=['POST'])
@login_required
def clear_tasks():
    username = session['username']

    if username in user_tasks:
        user_tasks[username] = []

    return jsonify({'status': 'success'})

@app.route('/rastreamento', methods=['GET', 'POST'])
@login_required
def rastreio():
    return render_template('sro.html')


import matplotlib.pyplot as plt
import pandas as pd


# @app.route('/dashboard', methods=['GET', 'POST'])
# @login_required
# def dashboard():
#     return render_template('teste.html')

@app.route('/add_note', methods=['POST'])
@login_required
def add_note():
    username = session['username']
    note_date_str = request.form.get('date')
    note_content = request.form.get('notes').replace('\n', '<br>')

    if note_date_str:
        
        note_date = datetime.strptime(note_date_str, '%d-%m-%Y').strftime('%Y-%m-%d')

        new_note = Note(username=username, date=note_date, content=note_content)
        db.session.add(new_note)
        db.session.commit()

        return jsonify({'status': 'success'})
    else:
        return jsonify({'status': 'error', 'message': 'Data inválida'})

@app.route('/get_notes', methods=['GET'])
@login_required
def get_notes():
    username = session['username']
    notes = Note.query.filter_by(username=username).order_by(Note.date).all()

    grouped_notes = defaultdict(list)
    for note in notes:
        if note.date:
            try:
                # Convertendo a data de volta para o formato d/m/Y
                note_date = datetime.strptime(note.date, '%Y-%m-%d')

                note_date_formatted = note_date.strftime('%d/%m/%Y') + ' - ' + note_date.strftime('%A')
                grouped_notes[note_date_formatted].append(note)
            except ValueError:

                pass

    return render_template('notes.html', grouped_notes=grouped_notes)

@app.route('/remove_note', methods=['POST'])
@login_required
def remove_note():
    note_id = request.form.get('note_id')
    username = session['username']

    note = Note.query.filter_by(id=note_id, username=username).first()

    if note:
        db.session.delete(note)
        db.session.commit()

        return redirect(url_for('get_notes'))
    else:
        return jsonify({'status': 'error', 'message': 'Nota não encontrada ou você não tem permissão para removê-la.'})
    
@app.route('/clear_notes', methods=['POST'])
@login_required
def clear_notes():
    username = session['username']

    # Remove todas as notas do usuário atual do banco de dados
    num_deleted = Note.query.filter_by(username=username).delete()
    db.session.commit()

    return redirect(url_for('get_notes'))

@app.route('/cadastro')
@login_required
def cadastro():
    return render_template('cadastro.html')

@app.route('/pesquisar', methods=['POST'])
@login_required
def pesquisar():
    nome_cliente = request.form['cliente']
    data_inicial = datetime.strptime(request.form['data_inicial'], '%Y-%m-%d').strftime('%d/%m/%Y')
    data_final = datetime.strptime(request.form['data_final'], '%Y-%m-%d').strftime('%d/%m/%Y')

    conn = sqlite3.connect('clientes.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM clientes WHERE cliente = ?", (nome_cliente,))
    resultado = cursor.fetchone()

    if resultado:
        token = resultado[2]
        conn.close()

        url = "https://api.boxlink.com.br/preenvio/consultar-periodo"
        payload = json.dumps({
            "dataInicial": request.form['data_inicial'],
            "dataFinal": request.form['data_final'],
            "preenvioCancelado": True,
            "envioExpedido": True
        })
        headers = {
            'Authorization': 'Bearer ' + token,
            'Content-Type': 'application/json'
        }
        response = requests.request("POST", url, headers=headers, data=payload)
        data = json.loads(response.text)
        chave_seller = sum('chaveSeller' in d for d in data)
        preEnvio_cancelado = sum(d['preenvioCancelado'] for d in data if 'preenvioCancelado' in d)
        envio_expedido = sum(d['envioExpedido'] for d in data if 'envioExpedido' in d)

        resultado_final = chave_seller - envio_expedido - preEnvio_cancelado

        mensagem = "{}\n\n de {} a {}:\n\n {} etiqueta(s) aguardando impressão".format(nome_cliente, data_inicial, data_final, resultado_final)
    else:
        mensagem = "Cliente não encontrado ou token não disponível."

    print("Mensagem retornada:", mensagem)

    return jsonify({'mensagem': mensagem})

@app.route('/adicionar', methods=['POST'])
@login_required
def adicionar():
    id = request.form['id']
    cliente = request.form['cliente']
    token = request.form['token']

    conn = sqlite3.connect('clientes.db')
    cursor = conn.cursor()
    cursor.execute('''INSERT INTO clientes (id, cliente, token)
                    VALUES (?, ?, ?)''', (id, cliente, token))
    conn.commit()
    conn.close()

    return redirect(url_for('cadastro'))

@app.route('/upload_csv', methods=['POST'])
@login_required
def upload_csv():
    if request.method == 'POST':

        if 'file' not in request.files:
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '' or not file.filename.endswith('.csv'):
            return redirect(request.url)
        
        try:
            process_csv(file)
            mensagem = "Dados adicionados com sucesso."
            return redirect(url_for('cadastro', mensagem=mensagem))
        except Exception as e:
            print(e)
            mensagem = "Ocorreu um erro ao processar o arquivo CSV."
            return redirect(url_for('cadastro', mensagem=mensagem))

def obter_dados_da_api(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        return None 

async def consulta_api_box_async(token, data_inicio, data_fim):
    nova_data_inicio = (datetime.strptime(data_inicio, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    url_api_box = "https://api.boxlink.com.br/preenvio/consultar-periodo"
    payload = {
        "dataInicial": nova_data_inicio,
        "dataFinal": data_fim,
        "preenvioCancelado": True,
        "envioExpedido": True
    }
    headers = {'Authorization': 'Bearer ' + token}
    
    for attempt in range(3): 
        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(url_api_box, json=payload, headers=headers, timeout=30) as response:
                    content_type = response.headers.get('Content-Type', '')
                    if 'application/json' in content_type:
                        data = await response.json()
                        
                        if isinstance(data, list) and all(isinstance(d, dict) for d in data):
                            chave_seller = sum('chaveSeller' in d for d in data)
                            preEnvio_cancelado = sum(d.get('preenvioCancelado', False) for d in data)
                            envio_expedido = sum(d.get('envioExpedido', False) for d in data)
                            resultado = chave_seller - envio_expedido - preEnvio_cancelado
                            return resultado
                        else:
                            print(f"Erro: 'data' retornado pela API não está no formato esperado. Data: {data}")
                            return None
                    else:
                        text = await response.text()
                        raise aiohttp.client_exceptions.ContentTypeError(
                            response.request_info,
                            response.history,
                            status=response.status,
                            message=f'Attempt to decode JSON with unexpected mimetype: {content_type}',
                            headers=response.headers,
                        )
        except (aiohttp.ClientError, asyncio.TimeoutError) as e:
            if attempt == 2:
                raise e
            await asyncio.sleep(2)

async def consulta_api_box(tokens, data_inicio, data_fim):
    tasks = []
    for token in tokens:
        task = consulta_api_box_async(token, data_inicio, data_fim)
        tasks.append(task)
    return await asyncio.gather(*tasks)

@app.route('/lista_completa')
async def lista_completa():
    # print("Request args:", request.args)
    data_inicio = request.args.get('data_inicio', '')
    data_fim = request.args.get('data_fim', '')

    if not data_inicio or not data_fim:
        return render_template('testeapi.html')
    
    try:
        data_inicio_formatted = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%Y%m%d')
        data_fim_formatted = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%Y%m%d')
    except ValueError:
        return "Formato de data inválido. Por favor, insira as datas no formato correto."

    try:
        url_api_crm = f"https://ap5tntnr6b.execute-api.us-east-1.amazonaws.com/api/transcrmc/{data_inicio_formatted}/{data_fim_formatted}"
        dados_api_crm = obter_dados_da_api(url_api_crm)
        # print("Dados retornados pela API CRM:", dados_api_crm)

        if dados_api_crm is None:
            return "Erro ao obter dados da primeira API"
        
        if not isinstance(dados_api_crm, list):
            # print(f"Erro: dados_api_crm deveria ser uma lista, mas é {type(dados_api_crm)}")
            return "Erro ao obter dados da primeira API"

        search_query = request.args.get('search', '')
        if search_query:
            dados_api_crm = [
                item for item in dados_api_crm if isinstance(item, dict) and search_query.lower() in item.get('Cliente', '').lower()
            ]

        dados_formatados = [(item['Id_Cliente'], item['Cliente'], item['Pedidos'], item['Token']) for item in dados_api_crm]
        # print("Dados formatados:", dados_formatados)

        tokens = [item['Token'] for item in dados_api_crm]
        try:
            resultado_final = await consulta_api_box(tokens, data_inicio, data_fim)
        except aiohttp.client_exceptions.ContentTypeError as e:
            print("Erro ao obter dados da segunda API:", e)
            return "Erro ao obter dados da segunda API"

        return render_template('testeapi.html', dados=dados_formatados, data_inicio=data_inicio, data_fim=data_fim, resultado_final=resultado_final)

    except Exception as e:
        print(f"Erro inesperado: {e}")
        return "Erro interno no servidor"


@app.route('/download', methods=['POST'])
def download():
    try:
        data = request.form.get('data')

        dados = eval(data)
        resultado_final = [row[3] for row in dados]

        df = pd.DataFrame(dados, columns=['Id_Cliente', 'Cliente', 'Pedidos CRM', 'Pedidos BOX'])

        output = BytesIO()

        df.to_excel(output, index=False)

        output.seek(0)

        response = make_response(send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
        response.headers["Content-Disposition"] = "attachment; filename=dados.xlsx"
        return response
    except Exception as e:
        return f"Erro ao processar o download: {str(e)}"

@app.route('/carregar_lista_responsaveis', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'Nenhum arquivo encontrado'
        
        file = request.files['file']
        
        if file.filename == '' or not file.filename.endswith('.xlsx'):
            return 'Por favor, selecione um arquivo Excel (.xlsx)'
        
        try:
            from io import BytesIO
            import pandas as pd
            
            criar_banco_dados()
            
            df = pd.read_excel(BytesIO(file.read()))
            
            inserir_dados_da_planilha(df)
            
            return 'Dados enviados com sucesso'
        except Exception as e:
            return f'Ocorreu um erro ao processar o arquivo: {str(e)}'
    
    return render_template('upload.html')

@app.route('/up_clientes')
@login_required
def up_clientes():
    return render_template('upload.html')

class Trello(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    responsavel = db.Column(db.String(100), nullable=False)
    CD = db.Column(db.String(100), nullable=False)
    link = db.Column(db.String(200), nullable=False)

class Links(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(100), nullable=False)
    url = db.Column(db.String(200), nullable=False)

@app.route('/links_uteis')
@login_required
def links():
    
    conn = sqlite3.connect('database1.db')
    cursor = conn.cursor()
    
    cursor.execute("SELECT nome, responsavel, CD, link FROM trello")
    clientes = cursor.fetchall()
    
    cursor.execute("SELECT titulo, url FROM links")
    outros_links = cursor.fetchall()
    
    conn.close()
    
    return render_template('links_uteis.html', clientes=clientes, outros_links=outros_links)

def connect_db():
    return sqlite3.connect('database1.db')

@app.route('/inserir_link', methods=['POST'])
@login_required
def inserir_link():
    if request.method == 'POST':
        titulo = request.form['titulo']
        url = request.form['url']

        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO links (titulo, url) VALUES (?, ?)", (titulo, url))
        conn.commit()
        conn.close()
        return redirect('/links_uteis')

@app.route('/inserir_trello', methods=['POST'])
@login_required
def inserir_trello():
    if request.method == 'POST':
        nome = request.form['nome']
        responsavel = request.form['responsavel']
        CD = request.form['CD']
        link = request.form['link']
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("INSERT INTO trello (nome, responsavel, CD, link) VALUES (?, ?, ?, ?)", (nome, responsavel, CD, link))
        conn.commit()
        conn.close()
        return redirect('/links_uteis')
    

@app.route('/buscacep', methods=['GET', 'POST'])
def buscacep():
    if request.method == 'POST':
        uf = request.form.get('uf')
        cidade = request.form.get('cidade')
        bairro = request.form.get('bairro')
        logradouro = request.form.get('logradouro')
        
        api_url = f"https://viacep.com.br/ws/{uf}/{cidade}/{logradouro}/json/"
        
        try:
            response = requests.get(api_url)
            response.raise_for_status()  # Verifica se ocorreu algum erro HTTP
            
            data = response.json()
            if data:  # Verificar se data não está vazio
                return render_template('buscacep.html', data=data)
            else:
                error = "Nenhum Resultado encontrado. Verifique os dados e tente novamente."
                return render_template('buscacep.html', error=error)
        
        except requests.exceptions.ConnectionError:
            error = "Erro de conexão. Verifique sua rede e tente novamente."
            return render_template('buscacep.html', error=error)
        
        except requests.exceptions.Timeout:
            error = "A requisição para a API expirou. Tente novamente mais tarde."
            return render_template('buscacep.html', error=error)
        
        except requests.exceptions.RequestException as e:
            error = f"Erro ao fazer a requisição: {e}"
            return render_template('buscacep.html', error=error)

    return render_template('buscacep.html')

@app.route('/buscacep_cep')
def bcep():
    return render_template('buscacep_cep.html')


# ROTAS E DEF FATURA EM PDF

# Função para quebrar texto
def wrap_text(text, width):
    return '\n'.join(textwrap.wrap(text, width=width))

# Função para converter DataFrame para lista de listas
def dataframe_to_list(dataframe):
    l = []
    lista = [dataframe.columns.values.tolist()] + dataframe.fillna('').values.tolist()
    del lista[0]
    l.append(["Objeto", "Postagem", "Valor", "Destinatário", "Cidade", "Observação"])
    for i in lista:
        try:
            observacao = i[11] if pd.notnull(i[11]) else ''
            l.append([
                i[0],
                i[3].strftime("%d/%m/%Y") if isinstance(i[3], datetime) else str(i[3]),
                ("R$ " + '%.2f' % i[4]).replace(".", ",") if isinstance(i[4], (int, float)) else str(i[4]),
                i[5],
                i[6],
                observacao
            ])
        except Exception as e:
            print(f"Erro ao processar linha: {i}, erro: {e}")
            continue
    return l

# Função para desenhar o cabeçalho
def header(canvas, doc, pini, pfin, cliente, estado):
    page_width, page_height = letter
    margin = 50
    canvas.saveState()
    canvas.setFont('Helvetica', 13)
    canvas.drawString(210, page_height - margin - 40, 'Fatura Resumida de Serviços Prestados')
    canvas.drawImage("static/correios.png", margin, page_height - margin - 40, width=120, height=25)
    canvas.setFont('Helvetica', 9)
    current_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    canvas.drawString(margin, page_height - margin - 75, f'Período de: {pini} até: {pfin} - Data da emissão: {current_date}')
    canvas.drawString(margin, page_height - margin - 85, f'Cliente: {cliente}')
    
    if estado == 'SC':
        canvas.drawString(margin, page_height - margin - 95, 'PALHOÇA - SC - PACHECOS - RUA VIKINGS 30 - 88134-878')
    elif estado == 'SP':
        canvas.drawString(margin, page_height - margin - 95, 'INDAIATUBA - SP - CENTRO - RUA ONZE DE JUNHO 1318 - 13330-972')
    elif estado == 'ES':
        canvas.drawString(margin, page_height - margin - 95, 'VITORIA - ES - CENTRO - RUA ENGENHEIRO PINTO PACCA 25 LJ D CAIXA POSTAL 10010  - 29010-973')
        
    canvas.setTitle(f"Relatórios Correios {pini} - {pfin}")
    canvas.restoreState()
    
# Função para desenhar o rodapé
def footer(canvas, doc):
    page_width, page_height = letter
    margin = 16
    
    canvas.saveState()
    
    rect_x = margin
    rect_y = margin
    rect_width = page_width - 2 * margin
    rect_height = 50
    radius = 10

    canvas.roundRect(rect_x, rect_y, rect_width, rect_height, radius, stroke=1, fill=0)

    canvas.setFont('Helvetica-Bold', 8.5)
    text = "A Empresa Brasileira de Correios e Telégrafos é imune a Imposto de Renda conforme sentença judicial - Recurso Extraordinário STF\n601.392/PR amparada no art.150, alínea a da CF/88 e no Decreto-lei 509/1969."
    text_lines = text.split('\n')
    
    text_x = rect_x + 10
    text_y = rect_y + rect_height - 15

    for line in text_lines:
        canvas.drawString(text_x, text_y, line)
        text_y -= 10 

    canvas.setFont('Helvetica-Oblique', 6)
    footer_text = "Para pagamento do Boleto junto ao seu Banco, se necessário utilize o CNPJ Matriz dos Correios: 34.028.316/0001-03 no campo Beneficiário, porexigência da CIP (Câmara interbancária de Pagamento)."
    
    text_y -= 20
    canvas.drawString(margin, text_y, footer_text)
    
    canvas.restoreState()

# Classe customizada para contar as páginas e adicionar o rodapé na última
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.pages = []

    def showPage(self):
        self.pages.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        page_count = len(self.pages)
        for page_number, page in enumerate(self.pages, start=1):
            self.__dict__.update(page)
            # Adicionar o rodapé apenas na última página
            if page_number == page_count:
                footer(self, None)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

def process_excel_to_pdf(file, pini, pfin, cliente, estado, nomearquivo):
    df = pd.read_excel(file, sheet_name="Planilha1")

    width = 20

    # Aplicar wrap_text apenas a colunas de texto
    df = df.applymap(lambda x: wrap_text(x, width) if isinstance(x, str) else x)

    data_list = dataframe_to_list(df)

    pdf_buffer = io.BytesIO()
    pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    elements = []

    table = Table(data_list, colWidths=[(612 - 2 * 50) / 6] * 6)
    bg_color = colors.Color(68 / 255, 114 / 255, 196 / 255)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), bg_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
        ('TOPPADDING', (0, 0), (-1, 0), 3),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0, colors.white)
    ])
    
    last_row = len(data_list) - 1
    style.add('BACKGROUND', (0, last_row), (-1, last_row), bg_color)
    style.add('TEXTCOLOR', (0, last_row), (-1, last_row), colors.white)

    table.setStyle(style)
    elements.append(table)

    # Adicionando o cabeçalho e o rodapé
    frame = Frame(pdf.leftMargin, pdf.bottomMargin, pdf.width, pdf.height - 2 * 50, id='normal')
    template = PageTemplate(id='test', frames=[frame], 
                            onPage=lambda canvas, doc: header(canvas, doc, pini, pfin, cliente, estado))
    pdf.addPageTemplates([template])

    # Construindo o PDF
    pdf.build(elements, canvasmaker=NumberedCanvas)

    # Retornando o buffer do PDF
    pdf_buffer.seek(0)
    return pdf_buffer

@app.route('/gerar_pdf', methods=['GET', 'POST'])
@login_required
def gerar_pdf():
    if request.method == 'POST':
        # Recebendo dados do formulário
        file = request.files['file']
        pini = request.form['pini']
        pfin = request.form['pfin']
        cliente = request.form['cliente']
        estado = request.form['estado']  # Captura o estado selecionado (SC ou SP)
        nomearquivo = request.form['nomearquivo']

        # Gerando o PDF
        pdf_buffer = process_excel_to_pdf(file, pini, pfin, cliente, estado, nomearquivo)

        # Enviando o PDF gerado para o cliente
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"{nomearquivo}.pdf",
            mimetype='application/pdf'
        )

    # Se for GET, apenas renderiza o formulário
    return render_template('gerar_pdf.html')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def normalizar_string(s):
    return s.upper().strip()

def somar_observacoes(observacao):
    produtos = {}
    if pd.isna(observacao):
        return produtos 

    observacao = str(observacao).strip()
    items = observacao.split('|')
    for item in items:
        item = item.strip()
        if 'x' in item:
            partes = item.split('x', 1)
            if len(partes) == 2:
                quantidade_str = partes[0].strip()
                produto = normalizar_string(partes[1].strip())
                
                try:
                    quantidade = int(quantidade_str)
                except ValueError:
                    quantidade = 1
                    
                if "EBOOK" not in produto:
                    if produto in produtos:
                        produtos[produto] += quantidade
                    else:
                        produtos[produto] = quantidade

    return produtos

@app.route('/somar_produtos')
def upload_file():
    return render_template('upload.html')

df_resultado = None

@app.route('/analisar', methods=['POST'])
def analisar():
    global df_resultado

    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']
    if file.filename == '':
        return redirect(request.url)

    if file:
        df = pd.read_csv(file, delimiter=';', encoding='ISO-8859-1')

        print("DataFrame Original:")
        print(df.head())

        df_observacoes = df['Observação'].apply(somar_observacoes)
        df_expanded = pd.json_normalize(df_observacoes)
        df = pd.concat([df, df_expanded], axis=1)

        df = df.drop(columns=['Observação'])

        print("DataFrame Após Expansão:")
        print(df.head())

        df_resultado = df.groupby('Data').sum().reset_index()

        print("DataFrame Após Agrupamento:")
        print(df_resultado.head())

        for coluna in df_resultado.columns:
            if df_resultado[coluna].dtype == 'float64':
                df_resultado[coluna] = df_resultado[coluna].fillna(0).astype(int)

        df_resultado['Data'] = pd.to_datetime(df_resultado['Data'], errors='coerce', format='%d/%m/%Y')
        df_resultado = df_resultado.sort_values(by='Data')

        df_resultado['Data'] = df_resultado['Data'].dt.strftime('%d/%m/%y')
        
        df_resultado = df_resultado.loc[:, ~df_resultado.columns.str.contains('^Unnamed')]

        print("DataFrame Após Remoção de Colunas 'Unnamed':")
        print(df_resultado.head())

        total_row = df_resultado.drop(columns=['Data']).sum()
        total_row['Data'] = 'Total Geral' 

        df_resultado.loc[len(df_resultado)] = total_row

        html_table = df_resultado.to_html(classes='table table-striped', index=False).strip()

        total_geral = df_resultado.drop(columns=['Data']).sum().sum()

        return render_template('resultado.html', 
                               tables=[html_table], 
                               titles=df_resultado.columns.values,
                               total_geral=total_geral)
        
@app.route('/download_xlsx')
def download_xlsx():
    global df_resultado

    if df_resultado is None:
        return redirect('/')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resultado.to_excel(writer, index=False, sheet_name='Resultado')

    output.seek(0)
    return send_file(output, as_attachment=True, download_name='resultado.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def processar_cancelamento(lines):
    url = "https://api.boxlink.com.br/v2/pre-envio/cancelar-com-rastreador"
    headers = {
        'Authorization': 'Bearer eyJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJqb25hc2dhcmNpYTY2NkBnbWFpbC5jb20iLCJVU0VSX0RFVEFJTFMiOnsidXNlcklkIjoxNTgzLCJtYXRyaXpJZCI6MTcsImZyYW5xdWlhSWQiOjksImNsaWVudGVJZCI6MjI1fSwiZXhwIjo1OTk1NzM4ODAwfQ.SEjKpWAkD_j5oosJ1RaSQq2JmMeXHhc459FqzJtxXc0',
        'User-Agent': 'Apidog/1.0.0 (https://apidog.com)',
        'Content-Type': 'application/json'
    }

    async def cancelar_rastreadores():
        async with aiohttp.ClientSession() as session:
            tasks = []
            for l in lines:
                payload = {
                    "rastreadorTms": l.strip(),
                    "motivo": "Solicitado pela Logistica"
                }
                task = asyncio.create_task(session.put(url, headers=headers, json=payload))
                tasks.append(task)
            responses = await asyncio.gather(*tasks)
            return responses

    return asyncio.run(cancelar_rastreadores())

@app.route('/cancelamento_etiquetas', methods=['GET', 'POST'])
@login_required
def cancelamento():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('cancelamento.html', message='Nenhum arquivo foi enviado')
        
        file = request.files['file']
        
        if file.filename == '':
            return render_template('cancelamento.html', message='Nenhum arquivo selecionado')
        
        if file and file.filename.endswith('.csv'):

            lines = file.read().decode('utf-8').splitlines()
            lines = lines[1:] 

            processar_cancelamento(lines)

            return render_template('cancelamento.html', message='Cancelamento realizado com sucesso')
        
        return render_template('cancelamento.html', message='Arquivo inválido. Por favor, revise a extensão (csv).')

    return render_template('cancelamento.html')

import httpx

async def fetch_news():
    api_key = "98e1265378d2abc81736b6495d5a0fe3"
    url = "https://gnews.io/api/v4/top-headlines"
    
    params = {
        "country": "br",
        "apikey": api_key,
        "lang": "pt",
        "max": 5
    }

    async with httpx.AsyncClient() as client:
        response = await client.get(url, params=params)
        
        if response.status_code == 200:
            articles = response.json().get("articles", [])
            return [
                {
                    "title": article["title"],
                    "description": article.get("description", "Sem descrição disponível"),
                    "url": article["url"],
                    "image": article.get("image", "path/to/placeholder.jpg")  # Adiciona o campo de imagem com um placeholder se não houver URL
                }
                for article in articles
            ]
        else:
            print("Erro ao buscar notícias:", response.status_code, response.text)
            return []

async def fetch_weather(city="Charqueadas", lat=None, lon=None):
    api_key = "eb8a5f8b0ace4c7fe4622f6deadcd5d0"
    url = "https://api.openweathermap.org/data/2.5/weather"
    
    params = {
        "appid": api_key,
        "units": "metric"
    }
    
    if lat and lon:
        params["lat"] = lat
        params["lon"] = lon
    else:
        params["q"] = city
    
    weather_translation = {
        "clear sky": "céu Limpo",
        "few clouds": "Poucas Nuvens",
        "scattered clouds": "Nuvens Dispersas",
        "broken clouds": "Nuvens Fragmentadas",
        "shower rain": "Chuva Leve",
        "rain": "Chuva",
        "thunderstorm": "Tempestade",
        "snow": "Neve",
        "mist": "Neblina",
        "overcast clouds": "Nublado",
        "moderate rain": "Chuva moderada",
    }
    
    async with httpx.AsyncClient() as client:
        response = await client.get(url, params=params)
        
        if response.status_code == 200:
            weather_data = response.json()
            description = weather_data['weather'][0]['description']
            translated_description = weather_translation.get(description, description)
            weather_data['weather'][0]['description'] = translated_description
            
            deg = weather_data['wind']['deg']
            directions = ['Norte', 'Nordeste', 'Leste', 'Sudeste', 'Sul', 'Sudoeste', 'Oeste', 'Noroeste']
            weather_data['wind_direction'] = directions[int((deg + 22.5) % 360 / 45)]
            
            return weather_data
        else:
            print("Erro ao buscar o clima:", response.status_code, response.text)
            return None

@app.route("/", methods=["GET", "POST"])
@login_required
def dashboard():
    lat = None
    lon = None
    if request.method == "POST":
        data = request.get_json()
        lat = data.get("lat")
        lon = data.get("lon")
        
        weather = asyncio.run(fetch_weather(lat=lat, lon=lon))
        return jsonify(weather) if weather else jsonify({"error": "Erro ao buscar clima"}), 500
    
    weather = asyncio.run(fetch_weather())
    news = asyncio.run(fetch_news())
    username = user_database.get(session.get('username', ''), {}).get('name', 'Convidado')
    
    return render_template("dashboard.html", weather=weather, news=news, username=username)

from google.oauth2.service_account import Credentials
import gspread

# Configurações
ESCOPOS = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]
PLANILHA_ID = "13Ivq0l0ueMB6GjO0xr6umLx7qHMvPJRAomjgxf3CunE"
CREDENCIAIS_JSON = os.getenv("GOOGLE_CREDENTIALS")  # Variável de ambiente para o JSON compactado

# Cabeçalho personalizado
CABECALHO_PERSONALIZADO = [
    "Data",
    "Nome da Empresa",
    "Email",
    "Nome dos Produtos",
    "Quando inicar os envios",
    "Pedidos em atraso",
    "Kits",
    "Modelo de Envio",
    "Acessos Plat. Vendas",
    "Acessos Plat. NFs",
    "CNPJ",
    "Fornecedor",
    "Data de Nascimento",
    "Info. ANVISA"
]


async def acessar_planilha_forms():
    try:
        if not CREDENCIAIS_JSON:
            raise Exception("Credenciais do Google não foram configuradas.")
        
        caminho_temp = "/tmp/credentials.json" 
        with open(caminho_temp, "w") as cred_file:
            cred_file.write(CREDENCIAIS_JSON)

        credenciais = Credentials.from_service_account_file(caminho_temp, scopes=ESCOPOS)
        cliente = gspread.authorize(credenciais)
        planilha = cliente.open_by_key(PLANILHA_ID)
        aba_forms = planilha.worksheet("Respostas ao formulário 1")

        dados_raw = aba_forms.get_all_records(empty2zero=False)
        colunas_planilha = aba_forms.row_values(1)

        dados_filtrados = [registro for registro in dados_raw if any(registro.values())]
        dados_formatados = []

        for registro in dados_filtrados:
            novo_registro = {}
            for i, coluna in enumerate(colunas_planilha):
                if coluna == "8 - Por gentileza nos encaminhe imagens dos produtos. (WHATSAPP)":
                    continue 
                chave = CABECALHO_PERSONALIZADO[i] if i < len(CABECALHO_PERSONALIZADO) else coluna
                novo_registro[chave] = registro.get(coluna, "")
            dados_formatados.append(novo_registro)

        return dados_formatados

    except gspread.exceptions.WorksheetNotFound:
        return {"erro": "A aba especificada não foi encontrada."}
    except Exception as e:
        return {"erro": f"Erro inesperado: {str(e)}"}


@app.route('/dados-forms')
@login_required
def exibir_dados():
    """Rota para exibir os dados da planilha."""
    dados = asyncio.run(acessar_planilha_forms())
    if "erro" in dados:
        return jsonify(dados)
    return render_template('dados_forms.html', dados=dados)

ALLOWED_EXTENSIONS1 = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file1(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS1

class Reverso(db.Model):
    __tablename__ = 'reverso'
    
    id = db.Column(db.Integer, primary_key=True)
    remetente = db.Column(db.String(100), nullable=False)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id', ondelete='SET NULL'), nullable=True)  # Mudando para nullable=True
    cod_rastreio = db.Column(db.String(100), nullable=False)
    descricao = db.Column(db.String(255), nullable=False)
    imagem = db.Column(db.String(255))
    criado_em = db.Column(db.DateTime, default=datetime.utcnow)

    cliente = db.relationship('Cliente', backref=db.backref('reversos', lazy=True, passive_deletes=True))
    
    def __repr__(self):
        return f'<Reverso {self.remetente} - {self.cod_rastreio}>'

class Cliente(db.Model):
    __tablename__ = 'cliente'

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    nome = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)

    def __repr__(self):
        return f'<Cliente {self.nome} - {self.email}>'

from flask_paginate import Pagination, get_page_parameter

@app.route('/reversos', methods=['GET'])
def reversos():
    query = request.args.get('q', '')  # Captura o termo de pesquisa
    start_date = request.args.get('start_date', '')  # Captura a data inicial
    end_date = request.args.get('end_date', '')  # Captura a data final
    page = request.args.get(get_page_parameter(), type=int, default=1)  # Pega a página atual
    per_page = 10  # Número de itens por página

    filters = []  # Lista para armazenar os filtros aplicados

    # Adiciona o filtro de pesquisa por nome do cliente, remetente ou código de rastreio
    if query:
        filters.append(
            or_(
                Cliente.nome.ilike(f'%{query}%'),  # Filtra pelo nome do cliente
                Reverso.remetente.ilike(f'%{query}%'),  # Filtra pelo remetente
                Reverso.cod_rastreio.ilike(f'%{query}%')  # Filtra pelo código de rastreio
            )
        )

    # Filtro de data de criação inicial
    if start_date:
        filters.append(Reverso.criado_em >= datetime.strptime(start_date, '%Y-%m-%d'))

    # Filtro de data de criação final
    if end_date:
        filters.append(Reverso.criado_em <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1))

    # Consulta os reversos aplicando os filtros
    reversos_query = Reverso.query.join(Cliente).filter(*filters).order_by(Reverso.criado_em.desc()).add_columns(
        Reverso.id,
        Cliente.nome.label('cliente'),
        Reverso.cod_rastreio.label('codigo'),
        Cliente.email.label('email'),
        Reverso.criado_em.label('data'),
        Reverso.remetente.label('remetente'),
        Reverso.descricao.label('descricao'),
        Reverso.imagem.label('imagem')
    )

    # Paginação dos resultados
    reversos = reversos_query.paginate(page=page, per_page=per_page, error_out=False)

    # Criação da navegação de páginas
    pagination = Pagination(page=page, total=reversos.total, per_page=per_page, record_name='reversos')

    # Retorna a resposta no formato JSON para requisições AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('partials/_reversos_list.html', reversos=reversos.items)

    # Caso não seja uma requisição AJAX, renderiza a página normal
    return render_template('listar_reversos.html', 
                           reversos=reversos.items, 
                           pagination=pagination, 
                           query=query,
                           start_date=start_date,
                           end_date=end_date)

@app.route('/reversos/exportar', methods=['GET'])
@login_required
def exportar_reversos():
    query = request.args.get('q', '')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    start_date = datetime.strptime(start_date, '%Y-%m-%d') if start_date else None
    end_date = datetime.strptime(end_date, '%Y-%m-%d') if end_date else None

    filters = []
    if query:
        filters.append(
            (Cliente.nome.contains(query) | Reverso.remetente.contains(query))
        )
    if start_date:
        filters.append(Reverso.criado_em >= start_date)
    if end_date:
        filters.append(Reverso.criado_em <= end_date)

    reversos_query = Reverso.query.join(Cliente).filter(*filters).add_columns(
        Cliente.nome.label('cliente'),
        Reverso.cod_rastreio.label('codigo'),
        Cliente.email.label('email'),
        Reverso.criado_em.label('data'),
        Reverso.remetente.label('remetente'),
        Reverso.descricao.label('descricao')
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reversos"
    
    ws.append(['Cliente', 'Codigo', 'Email', 'Data', 'Remetente', 'Descrição'])

    for reverso in reversos_query:
        ws.append([
            reverso.cliente,
            reverso.codigo,
            reverso.email,
            reverso.data.strftime('%d/%m/%Y'), 
            reverso.remetente,
            reverso.descricao
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"Reversos-{today}.xlsx"

    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return response

from flask import flash, redirect, url_for

br_tz = pytz.timezone('America/Sao_Paulo')

@app.before_request
def before_request():
    g.timezone = br_tz

@app.route('/reversos/adicionar', methods=['GET', 'POST'])
@login_required
def adicionar_reverso():
    if request.method == 'POST':
        remetente = request.form['remetente']
        cliente_id = request.form['cliente'] 
        cod_rastreio = request.form['cod_rastreio']
        descricao = request.form['descricao']
        imagem = request.files['imagem'] if 'imagem' in request.files else None

        agora = datetime.now(pytz.utc).astimezone(g.timezone)
        imagem_url = None
        imagem_bytes = None

        if imagem and allowed_file1(imagem.filename):
            imagem_bytes = imagem.read()

            filename = secure_filename(imagem.filename)

            s3 = boto3.client(
                's3',
                aws_access_key_id=app.config['AWS_ACCESS_KEY_ID'],
                aws_secret_access_key=app.config['AWS_SECRET_ACCESS_KEY'],
                region_name=app.config['AWS_S3_REGION_NAME']
            )

            try:
                s3.upload_fileobj(
                    BytesIO(imagem_bytes),
                    app.config['AWS_S3_BUCKET_NAME'],
                    f"uploads/{filename}",
                    ExtraArgs={'ContentType': imagem.content_type}
                )
                imagem_url = f"https://{app.config['AWS_S3_BUCKET_NAME']}.s3.{app.config['AWS_S3_REGION_NAME']}.amazonaws.com/uploads/{filename}"
            except NoCredentialsError:
                flash("Erro: Credenciais da AWS inválidas!", "danger")
                return redirect(url_for('adicionar_reverso'))
            except Exception as e:
                flash(f"Erro ao enviar a imagem ao S3: {e}", "danger")
                return redirect(url_for('adicionar_reverso'))

        cliente = Cliente.query.get(cliente_id)

        novo_reverso = Reverso(
            remetente=remetente,
            cliente=cliente,
            cod_rastreio=cod_rastreio,
            descricao=descricao,
            imagem=imagem_url,
            criado_em=agora
        )

        try:
            db.session.add(novo_reverso)
            db.session.commit()
        except Exception as e:
            flash(f"Erro ao salvar os dados do reverso: {e}", "danger")
            return redirect(url_for('adicionar_reverso'))

        try:
            msg = Message(
                subject=f"Logistica Reversa Recebida {cod_rastreio}",
                sender=app.config['MAIL_DEFAULT_SENDER'],
                recipients=[cliente.email],
                body=f"""
                Olá {cliente.nome},\n
                Recebemos uma devolução de Logística Reversa com os seguintes detalhes:
                - Remetente: {remetente}
                - Código de Rastreio: {cod_rastreio}
                
                - Descrição: {descricao}

                Atenciosamente,
                Equipe Conexão Premium
                """
            )

            if imagem_bytes:
                msg.attach(
                    filename=imagem.filename,
                    content_type=imagem.content_type,
                    data=imagem_bytes
                )

            mail.send(msg)
            flash("E-mail enviado com sucesso!", "success")
        except Exception as e:
            print(f"Erro ao enviar e-mail: {e}")
            flash(f"Erro ao enviar o e-mail: {str(e)}", "danger")

        return redirect(url_for('adicionar_reverso'))

    clientes = Cliente.query.all()
    return render_template('adicionar_reverso.html', clientes=clientes)


@app.route('/reversos/delete/<int:id>', methods=['GET'])
@login_required
def deletar_reverso(id):
    reverso = Reverso.query.get(id)

    if reverso:
        if reverso.imagem:
            try:
                s3 = boto3.client('s3')
                bucket_name = 'reversoscd-11cb1e80b53bd2a6b33fa34d587970b2'
                s3.delete_object(Bucket=bucket_name, Key=reverso.imagem)
                print(f"Imagem {reverso.imagem} excluída do S3.")
            except ClientError as e:
                print(f"Erro ao excluir a imagem do S3: {e}")
                flash("Erro ao excluir a imagem do S3.", "danger")
        
        try:
            db.session.delete(reverso)
            db.session.commit()
            flash("Reverso excluído com sucesso!", "success")
        except Exception as e:
            print(f"Erro ao excluir o reverso do banco de dados: {e}")
            flash("Erro ao excluir o reverso.", "danger")

    return redirect(url_for('reversos'))
    
@app.route('/cadastro_cliente', methods=['GET', 'POST'])
@login_required
def cadastro_cliente():
    if request.method == 'POST':
        nome = request.form['nome']
        email = request.form['email']
        
        if Cliente.query.filter_by(email=email).first():
            return "Este email já está cadastrado. Tente outro."
        
        novo_cliente = Cliente(nome=nome, email=email)
        db.session.add(novo_cliente)
        db.session.commit()
        
        return redirect(url_for('clientes')) 
    
    return render_template('clientes_reverso.html')

@app.route('/clientes')
@login_required
def clientes():
    clientes = Cliente.query.all()
    return render_template('clientes.html', clientes=clientes)

@app.route('/editar_cliente/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(id):
    cliente = Cliente.query.get_or_404(id)

    if request.method == 'POST':
        cliente.nome = request.form['nome']
        cliente.email = request.form['email']

        try:
            if Cliente.query.filter(Cliente.email == cliente.email, Cliente.id != id).first():
                flash("Este email já está em uso por outro cliente. Tente outro.", "warning")
                return redirect(url_for('editar_cliente', id=id))

            db.session.commit()
            flash("Cliente atualizado com sucesso!", "success")
            return redirect(url_for('clientes'))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao editar cliente: {e}", "danger")
    
    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/excluir_cliente/<int:id>', methods=['POST'])
@login_required
def excluir_cliente(id):
    cliente = Cliente.query.get_or_404(id)

    try:
        Reverso.query.filter_by(cliente_id=id).delete()
        db.session.commit()

        db.session.delete(cliente)
        db.session.commit()

        flash("Cliente e reversos excluídos com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir cliente: {e}", "danger")

    return redirect(url_for('clientes'))

@app.route('/analise-plataforma')
@login_required
def upload_page():
    return render_template('analise.html')

@app.route('/analise', methods=['POST'])
def analyze_file():
    if 'file' not in request.files:
        return "Nenhum arquivo enviado", 400

    file = request.files['file']
    if file.filename == '':
        return "Arquivo inválido", 400

    output = BytesIO()
    resultados, relatorio_detalhado = analyze_excel(file, output)
    output.seek(0)

    with open("resultado_analise.xlsx", "wb") as f:
        f.write(output.read())
    output.seek(0)

    with pd.ExcelWriter("relatorio_detalhado.xlsx", engine='openpyxl') as writer:
        relatorio_detalhado.to_excel(writer, index=False)

    return render_template('analise.html', result=resultados)

@app.route('/download-analise', methods=['GET'])
@login_required
def download_anl():
    output = "resultado_analise.xlsx"
    return send_file(output, as_attachment=True, download_name="resultado_analise.xlsx")

@app.route('/download-relatorio-detalhado', methods=['GET'])
@login_required
def download_relatorio_detalhado():
    output = "relatorio_detalhado.xlsx"
    return send_file(output, as_attachment=True, download_name="relatorio_detalhado.xlsx")

def analyze_excel(file, output):
    df = pd.read_excel(file, sheet_name="Envios Conexao")

    if "Identificador do Pedido" not in df.columns:
        raise ValueError("Coluna 'Identificador do Pedido' não encontrada na planilha.")

    pedidos = df["Identificador do Pedido"].astype(str).fillna("")

    categorias = [
        "Pedidos Manuais",
        "CoD",
        "Monetizze",
        "Reenvio",
        "Vazio",
        "Payt",
        "Braipp",
        "Outras Plataformas",
        "Nota Manual"
    ]

    contadores = {categoria: 0 for categoria in categorias}
    relatorio_detalhado = []

    for pedido in pedidos:
        sub_pedidos = pedido.split("|")
        for sub_pedido in sub_pedidos:
            sub_pedido = sub_pedido.strip()
            categoria = "Outras Plataformas"

            if sub_pedido == "":
                contadores["Vazio"] += 1
                categoria = "Vazio"
            elif sub_pedido.startswith("TmF"):
                contadores["Pedidos Manuais"] += 1
                categoria = "Pedidos Manuais"
            elif sub_pedido.startswith("ONDEM"):
                contadores["CoD"] += 1
                categoria = "CoD"
            elif sub_pedido.isdigit():
                contadores["Monetizze"] += 1
                categoria = "Monetizze"
            elif sub_pedido.startswith("Reenvio"):
                contadores["Reenvio"] += 1
                categoria = "Reenvio"
            elif re.fullmatch(r"[A-Za-z0-9]{6,7}|[A-Za-z]{6,7}", sub_pedido) and not (sub_pedido.startswith("M") and sub_pedido[1:].isdigit()):
                contadores["Payt"] += 1
                categoria = "Payt"
            elif sub_pedido.startswith("ven") and 3 < len(sub_pedido) <= 10:
                contadores["Braipp"] += 1
                categoria = "Braipp"
            elif sub_pedido.startswith("M") and sub_pedido[1:].isdigit():
                contadores["Nota Manual"] += 1
                categoria = "Nota Manual"

            relatorio_detalhado.append({"Identificador do Pedido": sub_pedido, "Plataforma": categoria})

    resultados_df = pd.DataFrame(list(contadores.items()), columns=["Categoria", "Quantidade"])
    relatorio_detalhado_df = pd.DataFrame(relatorio_detalhado)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        resultados_df.to_excel(writer, index=False)

    return contadores, relatorio_detalhado_df

import random
import string

def gerar_sequencia(tamanho=6):
    return f"REN{''.join(random.choices(string.ascii_uppercase + string.digits, k=tamanho))}"

def gerar_sequencia(tamanho=6):
    caracteres = string.ascii_uppercase + string.digits
    return f"REN{''.join(random.choices(caracteres, k=tamanho))}"

@app.route('/incluir_transacoes')
def incluir_transacoes():
    return render_template('codigo.html')

@app.route('/incluir', methods=['POST'])
def incluir():
    if 'file' not in request.files:
        flash("Nenhum arquivo enviado.")
        return redirect(url_for('incluir_transacoes'))

    file = request.files['file']
    if file.filename == '':
        flash("Nenhum arquivo selecionado.")
        return redirect(url_for('incluir_transacoes'))

    try:
        df = pd.read_csv(file, sep=';', engine='python', on_bad_lines='skip')
        df.columns = df.columns.str.strip()
    except Exception as e:
        flash(f"Erro ao ler o CSV: {e}")
        return redirect(url_for('incluir_transacoes'))

    if 'QUANTIDADE+PRODUTO' not in df.columns:
        flash("A coluna 'QUANTIDADE+PRODUTO' não foi encontrada.")
        return redirect(url_for('incluir_transacoes'))

    mascara = df['QUANTIDADE+PRODUTO'].notnull() & (df['QUANTIDADE+PRODUTO'].astype(str).str.strip() != '')
    df.loc[mascara, 'CONTEUDO'] = df.loc[mascara].apply(lambda _: gerar_sequencia(), axis=1)

    buffer = BytesIO()
    df.to_csv(buffer, sep=';', index=False)
    buffer.seek(0)

    return send_file(buffer,
                     as_attachment=True,
                     download_name='saida.csv',
                     mimetype='text/csv')
    
def get_tracking_info(cod_rastreio):
    url = f"https://nqvjhj9wef.execute-api.us-east-1.amazonaws.com/api/ar/{cod_rastreio}"
    try:
        response = requests.get(url)
        response.raise_for_status()  # Vai lançar uma exceção se o código de status não for 2xx
        return response.json()  # Retorna os dados da API em formato JSON
    except requests.exceptions.RequestException as e:
        return {'error': str(e)}

@app.route("/consulta-img", methods=["GET", "POST"])
@login_required
def track_package():
    if request.method == "POST":
        cod_rastreio = request.form["cod_rastreio"]
        tracking_info = get_tracking_info(cod_rastreio)
        
        nome_unidade, municipio, descricao_evento, imagem_base64, codigo = None, None, None, None, None
        
        if isinstance(tracking_info, list) and len(tracking_info) > 0:
            tracking_info = tracking_info[0]
            
            eventos = tracking_info.get("eventos", [])
            if eventos:
                evento = eventos[0]
                imagem_base64 = tracking_info.get("imagemBase64", None)
                nome_unidade = evento.get("nomeUnidade", "Desconhecido")
                municipio = evento.get("municipio", "Desconhecido")
                descricao_evento = evento.get("descricaoEvento", "Sem descrição")
                codigo = tracking_info.get("codigo", "Código não encontrado")
            else:
                nome_unidade, municipio, descricao_evento, imagem_base64 = "Sem eventos", "Sem eventos", "Sem eventos", None

        return render_template(
            "busca-img.html", 
            tracking_info=tracking_info,
            nome_unidade=nome_unidade,
            municipio=municipio,
            descricao_evento=descricao_evento,
            imagem_base64=imagem_base64,
            codigo=codigo
        )
    return render_template("busca-img.html", tracking_info=None)

import logging
from sqlalchemy.exc import OperationalError
from werkzeug.middleware.proxy_fix import ProxyFix

logger = logging.getLogger(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app)

# Modelo de Tarefa
class NewTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(50), default='To Do')
    assigned_to = db.Column(db.String(200), nullable=True)
    created_by = db.Column(db.String(100), nullable=False)

# Função para tentar commits com reconexão automática
def safe_commit():
    try:
        db.session.commit()
    except OperationalError:
        db.session.rollback()
        # logger.error("Erro de conexão com o banco, tentando novamente...")

# Rotas do Flask
@app.route('/trecco')
@login_required
def trecco():
    if 'username' not in session:
        return redirect(url_for('login'))

    tasks = NewTask.query.filter(
        (NewTask.assigned_to.like(f'%{session["username"]}%')) | 
        (NewTask.created_by == session['username'])
    ).all()

    archived_count = NewTask.query.filter(
        NewTask.status == 'Archived',
        (NewTask.assigned_to.like(f'%{session["username"]}%')) | 
        (NewTask.created_by == session['username'])
    ).count()

    return render_template('trecco.html', tasks=tasks, archived_count=archived_count)

@app.context_processor
def inject_todo_count():
    todo_count = 0
    if "username" in session:
        todo_count = NewTask.query.filter(
            NewTask.status == 'To Do',
            NewTask.assigned_to.like(f'%{session["username"]}%')
        ).count()
    return dict(todo_count=todo_count)

@app.route('/add', methods=['POST'])
@login_required
def add_task1():
    if 'username' not in session:
        return redirect(url_for('login'))

    title = request.form['title']
    description = request.form.get('description', '')
    assigned_to = request.form.getlist('assigned_to')

    if not assigned_to:
        return redirect(url_for('trecco'))

    for user in assigned_to:
        new_task = NewTask(
            title=title,
            description=description,
            assigned_to=user,
            created_by=session['username']
        )

        try:
            db.session.add(new_task)
            safe_commit()  # Chama a função que tenta o commit com reconexão
            logger.info(f'Tarefa adicionada com sucesso para {user}: {title}')
        except Exception as e:
            db.session.rollback()
            logger.error(f'Erro ao adicionar tarefa para {user}: {e}')
            return "Erro ao salvar a tarefa", 500

    return redirect(url_for('trecco'))

@app.route('/update/<int:task_id>', methods=['POST'])
@login_required
def update_task(task_id):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    task = NewTask.query.get(task_id)
    if task and task.assigned_to == session['username']:
        task.status = request.form['status']
        safe_commit()
    return redirect(url_for('trecco'))

@app.route('/delete/<int:task_id>', methods=['POST'])
@login_required
def delete_task(task_id):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    task = NewTask.query.get(task_id)
    if not task:
        flash("Tarefa não encontrada", "error")
        return redirect(url_for('trecco'))
    
    logger.info(f"Tarefa encontrada: {task.title}, Criada por: {task.created_by}")

    if task.created_by == session['username']:
        try:
            db.session.delete(task)
            db.session.commit()
            flash("Tarefa excluída com sucesso", "success")
        except Exception as e:
            db.session.rollback()
            flash("Erro ao excluir tarefa", "error")
            logger.error(f"Erro ao excluir tarefa: {e}")
    else:
        flash("Você não tem permissão para excluir essa tarefa", "error")

    return redirect(url_for('trecco'))

@app.route('/archive/<int:task_id>', methods=['POST'])
@login_required
def archive_task(task_id):
    if 'username' not in session:
        return redirect(url_for('login'))
    
    task = NewTask.query.get(task_id)
    if task and task.assigned_to == session['username']:
        task.status = 'Archived'
        safe_commit()
    return redirect(url_for('trecco'))

@app.route('/archived_tasks')
@login_required
def archived_tasks():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    archived_tasks = NewTask.query.filter(
        (NewTask.assigned_to == session['username']) | 
        (NewTask.created_by == session['username']),
        NewTask.status == 'Archived'
    ).all()
    
    return render_template('archived_tasks.html', tasks=archived_tasks)

@app.route('/add_task_form')
@login_required
def add_task_form():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    return render_template('add_task.html', users=user_database.keys(), user_database=user_database)

from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
from datetime import datetime, date




@app.route('/relatorio_reverso')
@login_required
def relatorio_reverso():
    """Página única com lista de relatórios e filtros"""
    try:
        # Parâmetros de filtro
        filtro_cliente = request.args.get('cliente', '').strip()
        data_inicial = request.args.get('data_inicial', '')
        data_final = request.args.get('data_final', '')
        buscar = request.args.get('buscar', '')
        
        # Inicializa variáveis
        relatorios_data = []
        relatorios = []
        clientes_unicos = set()
        erro = None
        
        # Só busca dados se o usuário clicou em buscar
        if buscar == 'true':
            # Busca relatórios da API
            relatorios_data = api_client.get_relatorio_completo(data_inicial or None, data_final or None)
        
            # Debug: mostra estrutura dos dados
            if relatorios_data:
                logger.info(f"Total de relatórios recebidos: {len(relatorios_data)}")
                if len(relatorios_data) > 0:
                    primeiro_relatorio = relatorios_data[0]
                    logger.info(f"Primeiro relatório - atributos: {dir(primeiro_relatorio)}")
                    if hasattr(primeiro_relatorio, 'raw_data'):
                        logger.info(f"Primeiro relatório - raw_data: {primeiro_relatorio.raw_data}")
            else:
                logger.warning("Nenhum relatório recebido da API")
            
            # Processa e filtra os dados
            for relatorio_data in relatorios_data:
                # Extrai dados do relatório - primeiro verifica raw_data
                raw_data = getattr(relatorio_data, 'raw_data', None)
                
                # Se tem raw_data, usa ele para extrair informações
                if raw_data and isinstance(raw_data, dict):
                    # Mapeia os campos corretos baseado na estrutura real da API
                    cliente = raw_data.get('Cliente', 'Cliente não informado')
                    data_reverso = raw_data.get('Data_Reverso')
                    codigo_reverso = raw_data.get('CodigoReverso')
                else:
                    # Fallback para atributos diretos
                    cliente = getattr(relatorio_data, 'Cliente', 'Cliente não informado')
                    data_reverso = getattr(relatorio_data, 'Data_Reverso', None)
                    codigo_reverso = getattr(relatorio_data, 'CodigoReverso', None)
                
                # Adiciona à lista de clientes únicos
                if cliente and cliente != 'Cliente não informado':
                    clientes_unicos.add(cliente)
                
                # Aplica filtro por cliente se especificado
                if filtro_cliente and filtro_cliente.lower() not in cliente.lower():
                    continue
                
                # Converte data se necessário
                if isinstance(data_reverso, str):
                    try:
                        from datetime import datetime
                        # Tenta diferentes formatos de data
                        for fmt in ['%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                            try:
                                data_reverso = datetime.strptime(data_reverso, fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            data_reverso = None
                    except:
                        data_reverso = None
                
                # Cria objeto simplificado
                relatorio = {
                    'id': getattr(relatorio_data, 'id', None),
                    'cliente': cliente,
                    'data_reverso': data_reverso.strftime('%d/%m/%Y') if data_reverso else 'Data não informada',
                    'data_reverso_raw': data_reverso.isoformat() if data_reverso else None,
                    'codigo_reverso': codigo_reverso,
                    'raw_data': raw_data
                }
                
                relatorios.append(relatorio)
            
            # Ordena por data (mais recente primeiro)
            relatorios.sort(key=lambda x: x['data_reverso_raw'] or '', reverse=True)
        
        return render_template('relatorio_reverso.html', 
                             relatorios=relatorios,
                             clientes_unicos=sorted(clientes_unicos),
                             filtro_cliente=filtro_cliente,
                             data_inicial=data_inicial,
                             data_final=data_final,
                             total_registros=len(relatorios),
                             erro=erro)
    
    except Exception as e:
        logger.error(f"Erro ao buscar relatórios: {e}")
        return render_template('relatorio_reverso.html', 
                             relatorios=[],
                             clientes_unicos=[],
                             filtro_cliente='',
                             data_inicial='',
                             data_final='',
                             total_registros=0,
                             erro=str(e))


@app.route('/api/relatorios')
def api_relatorios():
    """API endpoint para retornar dados em JSON"""
    try:
        filtro_cliente = request.args.get('cliente', '').strip()
        data_inicial = request.args.get('data_inicial', '')
        data_final = request.args.get('data_final', '')
        
        relatorios_data = api_client.get_relatorio_completo(data_inicial or None, data_final or None)
        
        # Processa os dados
        relatorios = []
        for relatorio_data in relatorios_data:
            cliente = getattr(relatorio_data, 'nome_fundo', None) or getattr(relatorio_data, 'codigo_fundo', None) or 'Cliente não informado'
            data_reverso = getattr(relatorio_data, 'data_relatorio', None) or getattr(relatorio_data, 'data_base', None)
            
            # Aplica filtro por cliente se especificado
            if filtro_cliente and filtro_cliente.lower() not in cliente.lower():
                continue
            
            relatorio = {
                'id': getattr(relatorio_data, 'id', None),
                'cliente': cliente,
                'data_reverso': data_reverso.isoformat() if data_reverso else None,
                'codigo_fundo': getattr(relatorio_data, 'codigo_fundo', None),
                'cnpj': getattr(relatorio_data, 'cnpj', None),
                'patrimonio_liquido': getattr(relatorio_data, 'patrimonio_liquido', None),
                'ativo_total': getattr(relatorio_data, 'ativo_total', None),
                'passivo_total': getattr(relatorio_data, 'passivo_total', None),
                'resultado_liquido': getattr(relatorio_data, 'resultado_liquido', None)
            }
            
            relatorios.append(relatorio)
        
        return jsonify({
            'sucesso': True,
            'total_registros': len(relatorios),
            'dados': relatorios,
            'timestamp': datetime.now().isoformat()
        })
    
    except Exception as e:
        logger.error(f"Erro na API: {e}")
        return jsonify({
            'sucesso': False,
            'erro': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500


@app.route('/debug')
def debug():
    """Página de debug para mostrar dados brutos da API"""
    try:
        relatorios_data = api_client.get_relatorio_completo()
        
        debug_info = {
            'total_relatorios': len(relatorios_data),
            'estrutura_primeiro': None,
            'dados_brutos': []
        }
        
        if relatorios_data:
            primeiro = relatorios_data[0]
            debug_info['estrutura_primeiro'] = {
                'atributos': [attr for attr in dir(primeiro) if not attr.startswith('_')],
                'raw_data': getattr(primeiro, 'raw_data', None)
            }
            
            # Mostra dados brutos dos primeiros 3 relatórios
            for i, rel in enumerate(relatorios_data[:3]):
                debug_info['dados_brutos'].append({
                    'indice': i,
                    'atributos': {attr: getattr(rel, attr, None) for attr in ['id', 'nome_fundo', 'codigo_fundo', 'data_relatorio', 'data_base']},
                    'raw_data': getattr(rel, 'raw_data', None)
                })
        
        return f"""
        <html>
        <head><title>Debug - Dados da API</title></head>
        <body>
            <h1>Debug - Dados da API</h1>
            <h2>Resumo</h2>
            <p>Total de relatórios: {debug_info['total_relatorios']}</p>
            
            <h2>Estrutura do Primeiro Relatório</h2>
            <pre>{debug_info['estrutura_primeiro']}</pre>
            
            <h2>Dados Brutos (primeiros 3)</h2>
            <pre>{debug_info['dados_brutos']}</pre>
            
            <p><a href="/">Voltar para página principal</a></p>
        </body>
        </html>
        """
    
    except Exception as e:
        return f"<h1>Erro no Debug</h1><p>{str(e)}</p><a href='/'>Voltar</a>"
    

SIMILARITY_THRESHOLD = 0.90

# Inicializa o sistema de aprendizado
learning_system = LearningSystem()

# ---------------- Similaridade ----------------
def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()

def normalize_text(text: str) -> str:
    text = text.lower().strip()
    return re.sub(r"\s+", " ", text)

# ---------------- Regras Manuais ----------------
def parse_manual_rules(rules_text: str) -> dict:
    """
    Converte:
    MAX VIGORAN PROMOCIONAL = Max Vigoran Black
    em:
    { 'max vigoran promocional': 'max vigoran black' }
    """
    rules = {}

    if not rules_text:
        return rules

    for line in rules_text.splitlines():
        if "=" not in line:
            continue

        left, right = line.split("=", 1)
        rules[normalize_text(left)] = normalize_text(right)

    return rules

def parse_manual_rules_with_originals(rules_text: str) -> tuple:
    """
    Retorna tanto as regras normalizadas quanto os originais para aprendizado
    Retorna: (rules_dict, originals_list)
    onde originals_list é [(original_left, original_right), ...]
    """
    rules = {}
    originals = []

    if not rules_text:
        return rules, originals

    for line in rules_text.splitlines():
        if "=" not in line:
            continue

        left, right = line.split("=", 1)
        left_original = left.strip()
        right_original = right.strip()
        
        # Armazena original e normalizado
        originals.append((left_original, right_original))
        rules[normalize_text(left_original)] = normalize_text(right_original)

    return rules, originals

def apply_manual_rules(product: str, rules: dict, learn: bool = False) -> str:
    """
    Aplica regra exata ou por contenção
    Nota: O aprendizado agora é feito antes do processamento, então learn=False por padrão
    """
    original_product = product
    for pattern, target in rules.items():
        if pattern in product:
            return target
    return original_product

# ---------------- Agrupamento ----------------
def group_products(products: dict) -> dict:
    grouped = {}

    for product, qty in products.items():
        target = None
        for group in grouped:
            if similarity(product, group) >= SIMILARITY_THRESHOLD:
                target = group
                break

        if target:
            grouped[target] += qty
        else:
            grouped[product] = qty

    return {k.title(): v for k, v in grouped.items()}

# ---------------- Parser ----------------
def parse_text(text: str, manual_rules: dict, generate_suggestions: bool = True) -> tuple:
    totals = defaultdict(int)
    suggestions_map = {}

    for line in text.splitlines():
        blocks = re.split(r"\s*\|\s*", line)

        for block in blocks:
            block = block.strip()
            if not block:
                continue

            # 🔹 NOVO: divide por "+"
            plus_blocks = re.split(r"\s*\+\s*", block)

            for part in plus_blocks:
                part = part.strip()
                if not part:
                    continue

                # Remove comentários/observações entre parênteses que vêm após o produto,
                # como "(ENVIAR SEDEX)" ou "(FALTOU PRODUTO)", para não atrapalhar o parser.
                # Ex.: "1 Necessarie (ENVIAR SEDEX) (FALTOU PRODUTO)" -> "1 Necessarie"
                part = re.sub(r"\([^)]*\)", "", part).strip()
                if not part:
                    continue

                # 🔹 Trata prefixo "n x ..." preservando contagens internas.
                #    - "5 x Be Libid"                  -> prefix_qty = 5 (se não houver número interno)
                #    - "1 x 3 Be Libid"                -> ignora o 1, usa o 3 interno
                #    - "1 x 3 HIDRALISO 1 SHAMPOO"     -> ignora o 1, captura os dois itens (3 Hidraliso, 1 Shampoo)
                prefix_qty = None
                m_prefix = re.match(r"^\s*(\d+)\s*x\s*(.+)$", part, flags=re.IGNORECASE)
                if m_prefix:
                    prefix_qty, rest = m_prefix.groups()
                    prefix_qty = int(prefix_qty)
                    rest = rest.strip()
                    # Se após o "x" já vem um número, mantemos esse número interno
                    # e ignoramos o prefixo (caso típico: "1 x 3 ...").
                    if re.match(r"^\d+\b", rest):
                        prefix_qty = None
                    part = rest

                # Encontra todos os produtos: com número ou sem número (quantidade 1)
                # Padrão: (número opcional) + nome do produto
                # O produto termina quando encontra outro número ou fim da string
                item_pattern = re.compile(
                    r"(?:(\d+)\s+)?([a-zà-ú][a-zà-ú0-9\s\-]+?)(?=\s*\d+\s+|$)",
                    flags=re.IGNORECASE
                )
                matches = item_pattern.findall(part)

                for qty, product in matches:
                    product = product.strip()
                    if not product:
                        continue
                    quantidade = int(qty) if qty else 1
                    if prefix_qty:
                        quantidade *= prefix_qty
                    original_product = product
                    product = normalize_text(product)
                    # Ignora "kit" e "kits"
                    if product in ['kit', 'kits']:
                        continue
                    
                    # Aplica regras manuais
                    product_before_rules = product
                    product = apply_manual_rules(product, manual_rules)
                    
                    # Se não houve transformação manual e queremos gerar sugestões
                    if generate_suggestions and product == product_before_rules:
                        if product not in suggestions_map:
                            suggestions = learning_system.suggest_transformation(product)
                            if suggestions:
                                # Verifica se há sugestão com 100% de confiança para aplicar automaticamente
                                high_confidence_suggestion = None
                                for sug in suggestions:
                                    if sug.get('confidence', 0) >= 1.0:  # 100% de confiança
                                        high_confidence_suggestion = sug
                                        break
                                
                                if high_confidence_suggestion:
                                    # Aplica automaticamente a transformação com 100% de confiança
                                    suggested_product = high_confidence_suggestion['suggested']
                                    product = normalize_text(suggested_product)
                                    
                                    # Aprende automaticamente com essa transformação
                                    learning_system.learn_rule(original_product, suggested_product)
                                    learning_system.record_success(original_product, suggested_product)
                                    
                                    # Registra que foi aplicado automaticamente (para exibir na interface)
                                    # Usa um dicionário para evitar duplicatas baseado em (original, suggested)
                                    if 'auto_applied' not in suggestions_map:
                                        suggestions_map['auto_applied'] = {}
                                    
                                    # Cria chave única baseada no original e sugerido (normalizados)
                                    auto_key = f"{normalize_text(original_product)}|{normalize_text(suggested_product)}"
                                    if auto_key not in suggestions_map['auto_applied']:
                                        suggestions_map['auto_applied'][auto_key] = {
                                            'original': original_product,
                                            'suggested': suggested_product,
                                            'confidence': 1.0
                                        }
                                else:
                                    # Adiciona sugestões normais (< 100%)
                                    suggestions_map[product] = suggestions

                    # Reaplica regras manuais após qualquer transformação automática,
                    # para permitir que regras manuais mais específicas sobrescrevam as automáticas.
                    product = apply_manual_rules(product, manual_rules)
                    
                    totals[product] += quantidade

    grouped = group_products(totals)
    return grouped, suggestions_map

def parse_csv(file, manual_rules: dict, generate_suggestions: bool = True) -> tuple:
    content = file.read().decode("utf-8", errors="ignore")
    return parse_text(content, manual_rules, generate_suggestions)

# ---------------- Rotas ----------------
@app.route("/saidas_diarias", methods=["GET"])
def saidas_diarias():
    return render_template("saidas_diarias.html")

@app.route("/processar", methods=["POST"])
def processar():
    rules_text = request.form.get("rules", "")
    manual_rules, originals = parse_manual_rules_with_originals(rules_text)
    
    # Aprende com todas as regras manuais fornecidas (usando textos originais)
    for original_pattern, original_target in originals:
        if original_pattern and original_target:
            try:
                rule_id = learning_system.learn_rule(original_pattern, original_target)
                print(f"Regra aprendida: '{original_pattern}' -> '{original_target}' (ID: {rule_id})")
            except Exception as e:
                print(f"Erro ao aprender regra '{original_pattern}' -> '{original_target}': {e}")

    if "file" in request.files and request.files["file"].filename:
        result, suggestions = parse_csv(request.files["file"], manual_rules)
    else:
        text = request.form.get("text", "")
        result, suggestions = parse_text(text, manual_rules)

    # Converte auto_applied de dicionário para lista (se existir)
    if suggestions and 'auto_applied' in suggestions:
        if isinstance(suggestions['auto_applied'], dict):
            suggestions['auto_applied'] = list(suggestions['auto_applied'].values())
    
    # Garante que suggestions seja sempre um objeto JSON válido
    suggestions_json = json.dumps(suggestions if suggestions else {}, ensure_ascii=False)
    
    return render_template(
        "resultado_saidas.html",
        result=result,
        result_json=json.dumps(result, ensure_ascii=False),
        rules=rules_text,
        suggestions=suggestions_json
    )

@app.route("/export", methods=["POST"])
def export_csv():
    result = json.loads(request.form["result"])

    # Cria um arquivo Excel em memória
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"

    # Cabeçalho
    ws.append(["Produto", "Quantidade"])

    # Dados
    for product, qty in result.items():
        ws.append([product, qty])

    # Salva em um buffer de bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resultado_produtos.xlsx"
    )

@app.route("/approve_suggestion", methods=["POST"])
def approve_suggestion():
    """Aprova uma sugestão e aprende com ela"""
    data = request.get_json()
    original = data.get("original", "").lower().strip()
    suggested = data.get("suggested", "").lower().strip()
    
    if original and suggested:
        learning_system.learn_rule(original, suggested)
        learning_system.record_success(original, suggested)
        return jsonify({"status": "success", "message": "Sugestão aprovada e aprendida!"})
    
    return jsonify({"status": "error", "message": "Dados inválidos"}), 400

@app.route("/reject_suggestion", methods=["POST"])
def reject_suggestion():
    """Rejeita uma sugestão"""
    data = request.get_json()
    original = data.get("original", "").lower().strip()
    suggested = data.get("suggested", "").lower().strip()
    
    if original and suggested:
        learning_system.record_failure(original, suggested)
        return jsonify({"status": "success", "message": "Sugestão rejeitada"})
    
    return jsonify({"status": "error", "message": "Dados inválidos"}), 400

@app.route("/learned_patterns", methods=["GET"])
def learned_patterns():
    """Retorna os padrões aprendidos"""
    limit = request.args.get("limit", 20, type=int)
    patterns = learning_system.get_learned_patterns(limit)
    return render_template("learned_patterns.html", patterns=patterns)

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
